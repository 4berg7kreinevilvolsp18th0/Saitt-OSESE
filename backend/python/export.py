"""
Data export functionality (CSV, Excel)
"""
import csv
import io
from typing import List, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from models import Appeal, Content, Direction


def export_appeals_to_csv(
    db: Session,
    appeals: List[Appeal],
    include_internal: bool = False
) -> str:
    """
    Export appeals to CSV format
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    headers = [
        "ID", "Заголовок", "Описание", "Статус", "Приоритет",
        "Направление", "Дата создания", "Дата закрытия",
        "Тип контакта", "Институт", "Категория"
    ]
    
    if include_internal:
        headers.extend(["Назначено", "Дедлайн", "Теги"])
    
    writer.writerow(headers)
    
    # Data rows
    for appeal in appeals:
        direction_name = ""
        if appeal.direction:
            direction_name = appeal.direction.title
        
        row = [
            str(appeal.id),
            appeal.title,
            appeal.description[:200] + "..." if len(appeal.description) > 200 else appeal.description,
            appeal.status,
            appeal.priority or "normal",
            direction_name,
            appeal.created_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.created_at else "",
            appeal.closed_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.closed_at else "",
            appeal.contact_type or "",
            appeal.institute or "",
            appeal.category or "",
        ]
        
        if include_internal:
            row.extend([
                str(appeal.assigned_to) if appeal.assigned_to else "",
                appeal.deadline.strftime("%Y-%m-%d") if appeal.deadline else "",
                ", ".join(appeal.tags) if appeal.tags else "",
            ])
        
        writer.writerow(row)
    
    return output.getvalue()


def export_appeals_to_excel(
    db: Session,
    appeals: List[Appeal],
    include_internal: bool = False
) -> bytes:
    """
    Export appeals to Excel format (requires openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"
    
    # Headers
    headers = [
        "ID", "Заголовок", "Описание", "Статус", "Приоритет",
        "Направление", "Дата создания", "Дата закрытия",
        "Тип контакта", "Институт", "Категория"
    ]
    
    if include_internal:
        headers.extend(["Назначено", "Дедлайн", "Теги"])
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_num, appeal in enumerate(appeals, 2):
        direction_name = ""
        if appeal.direction:
            direction_name = appeal.direction.title
        
        row = [
            str(appeal.id),
            appeal.title,
            appeal.description[:200] + "..." if len(appeal.description) > 200 else appeal.description,
            appeal.status,
            appeal.priority or "normal",
            direction_name,
            appeal.created_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.created_at else "",
            appeal.closed_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.closed_at else "",
            appeal.contact_type or "",
            appeal.institute or "",
            appeal.category or "",
        ]
        
        if include_internal:
            row.extend([
                str(appeal.assigned_to) if appeal.assigned_to else "",
                appeal.deadline.strftime("%Y-%m-%d") if appeal.deadline else "",
                ", ".join(appeal.tags) if appeal.tags else "",
            ])
        
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_statistics_to_csv(stats: dict) -> str:
    """
    Export statistics to CSV
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["Метрика", "Значение"])
    writer.writerow(["Всего обращений", stats.get("total", 0)])
    writer.writerow(["Создано сегодня", stats.get("created_today", 0)])
    writer.writerow(["Закрыто сегодня", stats.get("closed_today", 0)])
    writer.writerow([])
    writer.writerow(["По статусам"])
    for status, count in stats.get("by_status", {}).items():
        writer.writerow([status, count])
    writer.writerow([])
    writer.writerow(["По направлениям"])
    for direction_id, count in stats.get("by_direction", {}).items():
        writer.writerow([direction_id, count])
    
    return output.getvalue()

